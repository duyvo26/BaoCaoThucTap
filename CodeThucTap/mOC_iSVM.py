
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_protect
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
import io
import base64
import numpy as np
from datetime import datetime
import urllib
import shutil
from matplotlib import pyplot as plt
from CodeThucTap.static.ChucNang import Cut_file, find_best_nugamma
from CodeThucTap.static.mOC_iSVM_AP import mOC_iSVM_AP
from CodeThucTap.static.mOC_iSVM_nB import mOC_iSVM_nB
from CodeThucTap.static.mOC_iSVM_EP import mOC_iSVM_EP




def SaveFileDraw(request, name_, arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, numberCut, inputFor):
    import openpyxl
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = numberCut + 1
    column = inputFor + 1
    # Tạo một workbook mới và active nó
    wb = openpyxl.Workbook()
    ws = wb.active
    import string
    for ass in string.ascii_lowercase[:14]:
        ws.column_dimensions[ass].width = 25
    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    arrTam = []
    for i in range(0, row):

        if i == 0:
            arrTam.append("Batch Number")
        else:
            arrTam.append("Batch "+str(i))

    arr_out_ACC.insert(0, arrTam)
    arr_out_Precision.insert(0, arrTam)
    arr_out_Recall.insert(0, arrTam)
    arr_out_F1.insert(0, arrTam)

    # print(arr_out_ACC)
    for i in range(0, row):
        for j in range(0, column):
            # print(arr_out_ACC[j][i])
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_ACC[j][i]))

            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_ACC[j][i])

    # arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1
    for i in range(row, row * 2):
        for j in range(0, column):

            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_Precision[j][i - row]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_Precision[j][i - row])

    for i in range(row * 2, row * 3):
        for j in range(0, column):
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_Recall[j][i - (row * 2)]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_Recall[j][i - (row * 2)])
    for i in range(row * 3, row * 4):
        for j in range(0, column):
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_F1[j][i - (row * 3)]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_F1[j][i - (row * 3)])

    path_stock = settings.MEDIA_ROOT + \
        '/media/'+request.COOKIES.get('id')+'/'
    if (os.path.exists(path_stock+'/Excel') == False):
        os.mkdir(path_stock+'/Excel')
    now = datetime.now()
    dt_string = now.strftime("%d_%m_%Y___%H_%M_%S")
    output_excel_path = path_stock+'/Excel/' + \
        str(name_) + " " + str(dt_string)+'.xlsx'
    # Lưu lại file Excel
    wb.save(output_excel_path)
    return download(request, output_excel_path)


def download(request, file_path):

    # if os.path.exists(file_path):
    with open(file_path, 'rb') as fh:
        response = HttpResponse(
            fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + \
            os.path.basename(file_path)
        return response

def drawBieuDo(arr0, arr1, arr2, arr3, batch, name):
    arr_URL = []
    for sum_ in [arr0, arr1, arr2, arr3]:
        arr_for = []
        x_ve = np.arange(1, batch+1)
        for i in sum_:
            arr_for.append(float(i)*100)

        plt.plot(x_ve, arr_for, marker=5, linestyle='dashed',
                 label="mOC_iSVM."+name)
        plt.title("mOC_iSVM."+name)
        plt.axis([None, None, 0, 100])
        plt.legend()
        fig = plt.gcf()
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        buf.seek(0)
        string = base64.b64encode(buf.read())
        plt.close()
        uri = urllib.parse.quote(string)
        arr_URL.append(uri)

    array_SUM = []
    x_ve = np.arange(1, batch+1)
    for sum_ in [arr0, arr1, arr2, arr3]:
        array = []
        x_ve = np.arange(1, batch+1)
        for i in range(len(x_ve)):
            array.append(
                {"accuracy": (float(sum_[i])), "numbercut": x_ve[i]})
        array_SUM.append(array)
    return arr_URL, array_SUM



def get_random_string(length):
    import random
    import string
    # choose from all lowercase letter
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str


def SaveFileCut(request, ma, loai, data):
    path = settings.MEDIA_ROOT + '/media/' + \
        request.COOKIES.get('id')+'/CutFile/'
    namefile = ''
    if loai == 'partone':
        namefile = f"{ma}.train"
    elif loai == 'parttwo':
        namefile = f"{ma}.test"
    else:
        namefile = f"{ma}.data"

    print(namefile)
    if os.path.exists(path) == False:
        os.makedirs(path)
    # for i in data:
    f = open(path+namefile, "a+")
    f.writelines(data)
    # f.writelines('\n')
    f.close()
    


def cut_file(request, uploaded_file_url, filename, numberCut):
    dt_string = get_random_string(8)
    import pandas as pd
    data = pd.read_csv(uploaded_file_url)
    numberClass = 0
    for i in data.columns:
        if i == 'class':
            break
        numberClass += 1

    partone, parttwo, arr = Cut_file(
        numberClass, settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/'+os.sep+filename, int(numberCut))
    print("save file train")
    SaveFileCut(request, dt_string, 'partone', partone)
    print("save file test")
    SaveFileCut(request, dt_string, 'parttwo', parttwo)
    print("save file data")
    SaveFileCut(request, dt_string, 'arr', arr)
    path = settings.MEDIA_ROOT + '/media/' + \
        request.COOKIES.get('id')+'/CutFile/'

    return path+dt_string+'.train', path+dt_string+'.test', path+dt_string+'.data'
    # return render(request, 'svm_imoc.html', {'array': arr, 'part_two': parttwo, 'part_one': partone, "ma": dt_string})



def FindNuGamma(request):
    return render(request, "mOC_iSVM/FindNuGamma.html")


def FindNuGamma_POST(request):
    if request.method == 'POST':
        # cut file ##################################
        myfile_csv = request.FILES['file-csv-open']
        fs = FileSystemStorage(
            settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/')
        filename_csv = fs.save(myfile_csv.name, myfile_csv)
        uploaded_file_url = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/'+filename_csv
        filename_train, filename_test, filename_data = cut_file(
            request, uploaded_file_url, filename_csv, int(
                float(request.POST['input-pTC'])*10))
        # cut file ##################################

        # input
        nustart = float(request.POST['input-nu-start'])
        pTC = float(request.POST['input-pTC'])
        if float(request.POST['input-nu-start']) == 0:
            nustart = float(request.POST['input-nu-start']) + \
                float(request.POST['input-nu-step'])
        path = filename_data
        pathnuGa = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/find_best_nugamma'
        nuEnd = float(request.POST['input-nu-end'])
        inputNustep = float(request.POST['input-nu-step'])
        gamma_start = float(request.POST['input-gamma-start'])
        gammaEnd = float(request.POST['input-gamma-end'])
        gamma_step = float(request.POST['input-gamma-step'])
        # input end
        bestnu, bestgamma, acc, ArrNuGa = find_best_nugamma(
            pTC, path, pathnuGa, nustart, nuEnd, inputNustep, gamma_start, gammaEnd, gamma_step)

        shutil.rmtree(settings.MEDIA_ROOT+'/media/' +
                      request.COOKIES.get('id')+'/find_best_nugamma')
        print(f"{bestgamma} {bestnu}")
        return render(request, 'mOC_iSVM/FindNuGamma.html', {"bestnugamma": f"Best Nu Vừa Tìm được là:{bestnu} , Best Gamma Vừa tìm được là:{bestgamma} với acc la {acc} | acc , nu, gamma {ArrNuGa}"})


# AP
def One_mOC_iSVM_AP(request):
    return render(request, "mOC_iSVM/mOC_iSVM_AP.html")


def One_mOC_iSVM_AP_POST(request):
    if request.method == 'POST':
        # cut file ##################################
        myfile_csv = request.FILES['file-csv']
        fs = FileSystemStorage(
            settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/')
        filename_csv = fs.save(myfile_csv.name, myfile_csv)
        uploaded_file_url = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/'+filename_csv
        filename_train, filename_test, filename_data = cut_file(
            request, uploaded_file_url, filename_csv, int(
                request.POST['number-TrainTest']))
        # cut file ##################################
        age = int(request.POST['number-age'])
        batch = int(request.POST['number-Batch'])
        bestgamma = float(request.POST['number-best-gamma'])
        bestnu = float(request.POST['number-best-nu'])
        # chay 1 lan
        if str(request.POST['inputFor']) == '1':
            path = settings.MEDIA_ROOT+'/media/' + \
                request.COOKIES.get('id')+'/iSVM_AP/'+os.sep
            arr0, arr1, arr2, arr3 = mOC_iSVM_AP(request.COOKIES.get('id'),
                                                 filename_train, filename_test, path, bestnu, bestgamma, age, batch)
            arr_URL, array_SUM = drawBieuDo(
                arr0, arr1, arr2, arr3, batch, 'AP')
            return render(request, 'home/showData.html', {'img': arr_URL, "array_SUM": array_SUM, "drawchart": "is active"})
        else:
            # chay n lan
            arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1 = [], [], [], []
            for i in np.arange(1, int(request.POST['inputFor'])+1, 1):
                print("#########################################################")
                print(
                    f"###########        Da chay {i}               ###########")
                print("#########################################################")
                path = settings.MEDIA_ROOT+'/media/' + \
                    request.COOKIES.get('id')+'/iSVM_AP/'+os.sep
                arr0, arr1, arr2, arr3 = mOC_iSVM_AP(
                    filename_train, filename_test, path, bestnu, bestgamma, age, batch)

                arr0.insert(0, 'Accuracy Score lần chạy '+str(i))
                arr1.insert(0, 'Precision lần chạy '+str(i))
                arr2.insert(0, 'Recall lần chạy '+str(i))
                arr3.insert(0, 'F1 lần chạy '+str(i))

                arr_out_ACC.append(arr0)
                arr_out_Precision.append(arr1)
                arr_out_Recall.append(arr2)
                arr_out_F1.append(arr3)

            return SaveFileDraw(request, "AP", arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, int(request.POST['number-Batch']), int(request.POST['inputFor']))


# nB
def One_mOC_iSVM_nB(request):
    return render(request, "mOC_iSVM/mOC_iSVM_nB.html")


def One_mOC_iSVM_nB_POST(request):
    if request.method == 'POST':
        # cut file ##################################
        myfile_csv = request.FILES['file-csv']
        fs = FileSystemStorage(
            settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/')
        filename_csv = fs.save(myfile_csv.name, myfile_csv)
        uploaded_file_url = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/'+filename_csv
        filename_train, filename_test, filename_data = cut_file(
            request, uploaded_file_url, filename_csv, int(
                request.POST['number-TrainTest']))
        # cut file ##################################
        age = int(request.POST['number-age'])
        batch = int(request.POST['number-Batch'])
        bestgamma = float(request.POST['number-best-gamma'])
        bestnu = float(request.POST['number-best-nu'])
        # chay 1 lan
        if str(request.POST['inputFor']) == '1':
            path = settings.MEDIA_ROOT+'/media/' + \
                request.COOKIES.get('id')+'/iSVM_nB/'+os.sep
            arr0, arr1, arr2, arr3 = mOC_iSVM_nB(request.COOKIES.get('id'),
                                                 filename_train, filename_test, path, bestnu, bestgamma, age, batch)
            arr_URL, array_SUM = drawBieuDo(
                arr0, arr1, arr2, arr3, batch, 'nB')
            return render(request, 'home/showData.html', {'img': arr_URL, "array_SUM": array_SUM, "drawchart": "is active"})
        else:
            # chay n lan
            arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1 = [], [], [], []
            for i in np.arange(1, int(request.POST['inputFor'])+1, 1):
                print("#########################################################")
                print(
                    f"###########        Da chay {i}               ###########")
                print("#########################################################")
                path = settings.MEDIA_ROOT+'/media/' + \
                    request.COOKIES.get('id')+'/iSVM_nB/'+os.sep
                arr0, arr1, arr2, arr3 = mOC_iSVM_nB(request.COOKIES.get('id'),
                                                     filename_train, filename_test, path, bestnu, bestgamma, age, batch)

                arr0.insert(0, 'Accuracy Score lần chạy '+str(i))
                arr1.insert(0, 'Precision lần chạy '+str(i))
                arr2.insert(0, 'Recall lần chạy '+str(i))
                arr3.insert(0, 'F1 lần chạy '+str(i))

                arr_out_ACC.append(arr0)
                arr_out_Precision.append(arr1)
                arr_out_Recall.append(arr2)
                arr_out_F1.append(arr3)

            return SaveFileDraw(request, "nB", arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, int(request.POST['number-Batch']), int(request.POST['inputFor']))


# EP
def One_mOC_iSVM_EP(request):
    return render(request, "mOC_iSVM/mOC_iSVM_EP.html")


def One_mOC_iSVM_EP_POST(request):
    if request.method == 'POST':
        # cut file ##################################
        myfile_csv = request.FILES['file-csv']
        fs = FileSystemStorage(
            settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/')
        filename_csv = fs.save(myfile_csv.name, myfile_csv)
        uploaded_file_url = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/'+filename_csv
        filename_train, filename_test, filename_data = cut_file(
            request, uploaded_file_url, filename_csv, int(
                request.POST['number-TrainTest']))
        # cut file ##################################
        age = float(request.POST['number-age'])
        batch = int(request.POST['number-Batch'])
        bestgamma = float(request.POST['number-best-gamma'])
        bestnu = float(request.POST['number-best-nu'])
        # chay 1 lan
        if str(request.POST['inputFor']) == '1':
            path = settings.MEDIA_ROOT+'/media/' + \
                request.COOKIES.get('id')+'/iSVM_EP/'+os.sep
            arr0, arr1, arr2, arr3 = mOC_iSVM_EP(request.COOKIES.get('id'),
                                                 filename_train, filename_test, path, bestnu, bestgamma, age, batch)
            arr_URL, array_SUM = drawBieuDo(
                arr0, arr1, arr2, arr3, batch, 'EP')
            return render(request, 'home/showData.html', {'img': arr_URL, "array_SUM": array_SUM, "drawchart": "is active"})
        else:
            # chay n lan
            arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1 = [], [], [], []
            for i in np.arange(1, int(request.POST['inputFor'])+1, 1):
                print("#########################################################")
                print(
                    f"###########        Da chay {i}               ###########")
                print("#########################################################")
                path = settings.MEDIA_ROOT+'/media/' + \
                    request.COOKIES.get('id')+'/iSVM_EP/'+os.sep
                arr0, arr1, arr2, arr3 = mOC_iSVM_EP(request.COOKIES.get('id'),
                                                     filename_train, filename_test, path, bestnu, bestgamma, age, batch)

                arr0.insert(0, 'Accuracy Score lần chạy '+str(i))
                arr1.insert(0, 'Precision lần chạy '+str(i))
                arr2.insert(0, 'Recall lần chạy '+str(i))
                arr3.insert(0, 'F1 lần chạy '+str(i))

                arr_out_ACC.append(arr0)
                arr_out_Precision.append(arr1)
                arr_out_Recall.append(arr2)
                arr_out_F1.append(arr3)

            return SaveFileDraw(request, "EP", arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, int(request.POST['number-Batch']), int(request.POST['inputFor']))
