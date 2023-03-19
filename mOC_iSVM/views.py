from datetime import datetime
import io
import re
from unicodedata import name
from django.shortcuts import redirect, render
from django.http import HttpResponse, StreamingHttpResponse
import json
import os
import base64
import urllib
import os
from matplotlib.font_manager import json_dump
from numpy import array, partition
from django.views.decorators.csrf import csrf_protect
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from mOC_iSVM.static.CodeCoDien import Select_Model
from mOC_iSVM.static.ChucNang import ChuanDoanCoDien
from matplotlib import pyplot as plt
import shutil
import csv
import numpy as np
import os
import zipfile
import pandas as pd
import sys
import json
from django.http import HttpResponse


# dung chung
sys.stdin.reconfigure(encoding='utf-8')
sys.stdout.reconfigure(encoding='utf-8')


def get_random_string(length):
    import random
    import string
    # choose from all lowercase letter
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str


def SaveFileDraw(request, name_, arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, numberCut, inputFor):
    import openpyxl
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = numberCut + 1
    column = inputFor + 1
    # Tạo một workbook mới và active nó
    wb = openpyxl.Workbook()
    ws = wb.active
    import string
    for ass in string.ascii_lowercase[:14]:
        ws.column_dimensions[ass].width = 25
    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    arrTam = []
    for i in range(0, row):

        if i == 0:
            arrTam.append("Batch Number")
        else:
            arrTam.append("Batch "+str(i))

    arr_out_ACC.insert(0, arrTam)
    arr_out_Precision.insert(0, arrTam)
    arr_out_Recall.insert(0, arrTam)
    arr_out_F1.insert(0, arrTam)

    # print(arr_out_ACC)
    for i in range(0, row):
        for j in range(0, column):
            # print(arr_out_ACC[j][i])
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_ACC[j][i]))

            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_ACC[j][i])

    # arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1
    for i in range(row, row * 2):
        for j in range(0, column):

            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_Precision[j][i - row]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_Precision[j][i - row])

    for i in range(row * 2, row * 3):
        for j in range(0, column):
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_Recall[j][i - (row * 2)]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_Recall[j][i - (row * 2)])
    for i in range(row * 3, row * 4):
        for j in range(0, column):
            if i == 0:
                ws.cell(column=j+1, row=i+1,
                        value=str(arr_out_F1[j][i - (row * 3)]))
            else:
                ws.cell(column=j+1, row=i+1,
                        value=arr_out_F1[j][i - (row * 3)])

    path_stock = settings.MEDIA_ROOT + \
        '/media/'+request.COOKIES.get('id')+'/'
    if (os.path.exists(path_stock+'/Excel') == False):
        os.mkdir(path_stock+'/Excel')
    now = datetime.now()
    dt_string = now.strftime("%d_%m_%Y___%H_%M_%S")
    output_excel_path = path_stock+'/Excel/' + \
        str(name_) + " " + str(dt_string)+'.xlsx'
    # Lưu lại file Excel
    wb.save(output_excel_path)
    return download(request, output_excel_path)


def download(request, file_path):

    # if os.path.exists(file_path):
    with open(file_path, 'rb') as fh:
        response = HttpResponse(
            fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + \
            os.path.basename(file_path)
        return response


def exportcsv(request, ma, loai):
    if loai == 'part_one':
        namefile = f"{ma}.train"
    elif loai == 'part_two':
        namefile = f"{ma}.test"
    else:
        namefile = f"{ma}.data"
    path = settings.MEDIA_ROOT + '/media/' + \
        request.COOKIES.get('id')+'/CutFile/'+namefile

    # filename = obj.model_attribute_name.path
    from django.http import FileResponse
    response = FileResponse(open(path, 'rb'))
    return response


def listModel(request):
    try:
        from mOC_iSVM.static import ChucNang
        listModel = ChucNang.ShowModel(request.COOKIES.get('id'))
    except:
        listModel = ""

    return render(request, 'home/listModel.html', {"listModel": listModel})


def DowMOdel(request):
    fileName = settings.MEDIA_ROOT + 'media/' + \
        request.COOKIES.get('id') + "/DowModels/" + \
        request.POST['NameFile']

    print(fileName)
    return download(request, fileName)

# end dung chung #########################################################


# home du doan mo hinh va cai dat
def index(request):
    return render(request, 'home/home.html')

def load (request):
    return render(request, 'load.html')




def setting(request):
    fontSize = range(10, 55)
    return render(request, 'home/setting.html', {'fontSize': fontSize})

def caidat(request):
    fontSize = range(10, 55)
    return render(request, 'home/caidat.html', {'fontSize': fontSize})



def setting_POST(request):
    if request.method == 'POST':
        myfile, fileAnh = "", ''
        try:
            myfile = request.FILES['input-file']
            fs = FileSystemStorage(settings.MEDIA_ROOT +
                                '/media/'+request.COOKIES.get('id')+'/FileLoadModel/')
            filename = fs.save(myfile.name, myfile)
            myfile = myfile.name
        except:
            myfile = ''

        try:
            fileAnh = request.FILES['input-anhNen']
            fs = FileSystemStorage(settings.MEDIA_ROOT +
                                '/images/FileLoadAnhNen/')
            filename = fs.save(fileAnh.name, fileAnh)
            fileAnh = fileAnh.name
        except:
            fileAnh = ''
        
        
        from django.http import JsonResponse
        data = {
            'myfile': myfile,
            'fileAnh': fileAnh,
            
            'thongbao': 'Tải lên thành công'
        }

        return JsonResponse(data)




def showData(request):
    return render(request, 'home/showData.html')

# ve bieu do


def drawBieuDo(arr0, arr1, arr2, arr3, batch, name):
    arr_URL = []
    for sum_ in [arr0, arr1, arr2, arr3]:
        arr_for = []
        x_ve = np.arange(1, batch+1)
        for i in sum_:
            arr_for.append(float(i)*100)

        plt.plot(x_ve, arr_for, marker=5, linestyle='dashed',
                 label="mOC_iSVM."+name)
        plt.title("mOC_iSVM."+name)
        plt.axis([None, None, 0, 100])
        plt.legend()
        fig = plt.gcf()
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        buf.seek(0)
        string = base64.b64encode(buf.read())
        plt.close()
        uri = urllib.parse.quote(string)
        arr_URL.append(uri)

    array_SUM = []
    x_ve = np.arange(1, batch+1)
    for sum_ in [arr0, arr1, arr2, arr3]:
        array = []
        x_ve = np.arange(1, batch+1)
        for i in range(len(x_ve)):
            array.append(
                {"accuracy": (float(sum_[i])), "numbercut": x_ve[i]})
        array_SUM.append(array)
    return arr_URL, array_SUM


# end home du doan mo hinh va cai dat
def GiaiThuatCoDien_Show(request):
    return render(request, "OneFile/GiaiThuatCoDien_Show.html")



# one file
def One_GiaiThuatCoDien(request):
    return render(request, "OneFile/GiaiThuatCoDien.html")

# co dien


def One_GiaiThuatCoDien_POST(request):
    if request.method == 'POST':
        myfile = request.FILES['file-csv-open']
        fs = FileSystemStorage(settings.MEDIA_ROOT +
                               '/media/'+request.COOKIES.get('id')+'/')
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        FILECSV = settings.MEDIA_ROOT + "/media/" + \
            request.COOKIES.get('id') + "/" + myfile.name
        batch = int(request.POST['number-cut'])
        age = int(request.POST['number-age'])
        select = request.POST['select']
        path = settings.MEDIA_ROOT+'/media/'+request.COOKIES.get('id')+'/'
        dataPhu = request.POST['dataPhu']
        number_k = float(request.POST['number-k'])
        if str(request.POST['inputFor']) == '1':
            arr0, arr1, arr2, arr3 = Select_Model(request.COOKIES.get('id'),
                                                  dataPhu, number_k, FILECSV, path, batch, age, select)
            arr_URL, array_SUM = [], []
            for sum_ in [arr0, arr1, arr2, arr3]:
                arr_for = [float(i)*100 for i in sum_]
                x_ve = np.arange(1, batch+1)
                plt.plot(x_ve, arr_for, marker=5, linestyle='dashed',
                         label=request.POST['select'])
                plt.title(request.POST['select'])
                plt.axis([None, None, 0, 100])
                plt.legend()
                fig = plt.gcf()
                buf = io.BytesIO()
                fig.savefig(buf, format='png')
                buf.seek(0)
                string = base64.b64encode(buf.read())
                plt.close()
                uri = urllib.parse.quote(string)
                arr_URL.append(uri)
                array = [{"accuracy": (float(sum_[i])), "numbercut": x_ve[i]}
                         for i in range(len(x_ve))]
                array_SUM.append(array)

            return render(request, 'home/showData.html', {'img': arr_URL, "array_SUM": array_SUM, "draw": "is active"})
        else:
            arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1 = [], [], [], []
            for i in np.arange(1, int(request.POST['inputFor'])+1, 1):
                arr0, arr1, arr2, arr3 = Select_Model(request.COOKIES.get('id'),
                    dataPhu, number_k, FILECSV, path, batch, age, select)

                arr0.insert(0, 'Accuracy Score lần chạy '+str(i))
                arr1.insert(0, 'Precision lần chạy '+str(i))
                arr2.insert(0, 'Recall lần chạy '+str(i))
                arr3.insert(0, 'F1 lần chạy '+str(i))

                arr_out_ACC.append(arr0)
                arr_out_Precision.append(arr1)
                arr_out_Recall.append(arr2)
                arr_out_F1.append(arr3)

            return SaveFileDraw(request, f"{request.POST['select']} Classical algorithm", arr_out_ACC, arr_out_Precision, arr_out_Recall, arr_out_F1, int(request.POST['number-cut']), int(request.POST['inputFor']))

    else:
        return redirect('test')

# Two file



# Chuan doan

def ChuanDoanUploadFIle(request, FileCSV, FileModel, ListClassOut):

    ArrOUT = []
    ArrOUTKQ = ['Out Lable']
    with open(FileCSV, 'r', encoding="utf-8") as csvfile:
        csvreader = csv.reader(csvfile)
        header = next(csvreader)
        PointClass = header.index('class')
        ArrOUT.append(header)
        data = []
        for row in csvreader:
            try:
                row_float = [float(val) for val in row]
            except:
                row_float = [float(val) for val in row]
            data.append(row_float)

    for index, line in enumerate(data):
        arrCheck = []
        for ind, lie in enumerate(line):
            if ind == PointClass:
                continue
            arrCheck.append(lie)
        try:
            outClass, outPhanTram = ChuanDoanCoDien(FileModel, arrCheck)
        except Exception as es:
            print(es)
            return render(request, 'home/home.html', {'ketqua': "Có lỗi vui lòng thử lại !!!"})

        ArrOUT.append(line)
        ArrOUTKQ.append([str(ListClassOut[outClass[0]]), outClass[0]])

    path = settings.MEDIA_ROOT+'/media/' + \
        request.COOKIES.get('id')+'/'+str(get_random_string(8))+'_OutClass.csv'

    with open(path, 'w', newline='', encoding="utf-8") as csvfile:
        csvwriter = csv.writer(csvfile)
        for row in ArrOUT:
            csvwriter.writerow(row)

    with open(path, 'r', encoding="utf-8") as f:
        reader = csv.reader(f)
        data = list(reader)

    for index, line in enumerate(ArrOUTKQ):
        data[index].append(line)

    with open(path, 'w', newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(data)
    return download(request, path)


def ChuanDoan(request):
    if request.method == 'POST':
        # upload file len
        # input-file
        # myfile = request.FILES['input-file']
        # fs = FileSystemStorage(
        #     settings.MEDIA_ROOT+'/media/' + request.COOKIES.get('id')+'/tmp/UpModel/')
        # filename = fs.save(myfile.name, myfile)
        
        # settings.MEDIA_ROOT +
        #                         '/media/'+request.COOKIES.get('id')+'/FileLoadModel/
                                
        FileModel = settings.MEDIA_ROOT+'/media/' + \
            request.COOKIES.get('id')+'/FileLoadModel/' +  request.POST['FileMOdel']

        # select mo hinh du doan
        select = request.POST['select']
        if select == 'CoDien':
            textIn = []
            print("---------------------------")
            if request.POST['selectChuanDoan'] == 'UpCSV':
                myfile = request.FILES['input-fileCSV']
                fs = FileSystemStorage(
                    settings.MEDIA_ROOT+'/media/' + request.COOKIES.get('id')+'/tmp/UpModel/')
                filename = fs.save(myfile.name, myfile)
                FileCSV = settings.MEDIA_ROOT+'/media/' + \
                    request.COOKIES.get('id')+'/tmp/UpModel/' + myfile.name

                return ChuanDoanUploadFIle(request, FileCSV, FileModel, request.POST['ListNameClass'].split(','))
            if request.POST['selectChuanDoan'] == 'Chuoi':
                # chuoi text
                for i in request.POST['inputSum'].split(','):
                    textIn.append(float(i))
            if request.POST['selectChuanDoan'] == 'Input':
                # nhap tung o
                for i in range(0, int(request.POST['S_input'])):
                    thamSO = str(request.POST[f'input{i}']).replace(" ", '')
                    textIn.append(float(thamSO))

            outClass = 0
            try:
                outClass, outPhanTram = ChuanDoanCoDien(FileModel, textIn)
            except Exception as es:
                print(es)
                return render(request, 'home/home.html', {'ketqua': "Có lỗi vui lòng thử lại !!!"})

            print("ChuanDoanCoDien-----------------------------------------", outClass)

            try:
                getClassName = request.POST['ListNameClass'].split(',')[
                    outClass[0]]
            except:
                getClassName = f"Không tìm thấy tên Class, Class mặt định là: {outClass[0]}"

            out = f"Kết quả: {getClassName}. Độ chính xác {outPhanTram*100} %"

            return render(request, 'home/home.html', {'ketqua': out})

